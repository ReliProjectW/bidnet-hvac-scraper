#!/usr/bin/env python3
"""
Export Smart Contract Parser Results to Excel
Creates clean Excel file from the latest smart parsing results
"""

import pandas as pd
import subprocess
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

def play_alert(message="Task complete"):
    """Play terminal bell and system notification"""
    print("\a", end="", flush=True)
    try:
        subprocess.run(['osascript', '-e', f'display notification "{message}" with title "Excel Export"'], 
                      capture_output=True, check=False)
    except:
        pass

def parse_smart_contracts_file(filepath):
    """Parse the smart contracts text file into structured data"""
    contracts = []
    
    with open(filepath, 'r') as f:
        content = f.read()
    
    # Handle the raw newline characters in the file
    content = content.replace('\\n', '\n')
    
    # Split by contract sections - look for the pattern "Contract X [TYPE]:"
    import re
    contract_pattern = r'Contract \d+ \[[A-Z_]+\]:'
    contract_sections = re.split(contract_pattern, content)[1:]  # Skip the header part
    
    for section in contract_sections:
        lines = section.strip().split('\n')
        if not lines:
            continue
            
        contract = {}
        
        for line in lines:
            line = line.strip()
            if line.startswith('Title: '):
                contract['title'] = line.replace('Title: ', '').strip()
            elif line.startswith('Primary Agency: '):
                contract['primary_agency'] = line.replace('Primary Agency: ', '').strip()
            elif line.startswith('Secondary Agency: '):
                contract['secondary_agency'] = line.replace('Secondary Agency: ', '').strip()
            elif line.startswith('Location: '):
                contract['location'] = line.replace('Location: ', '').strip()
            elif line.startswith('Description: '):
                contract['description'] = line.replace('Description: ', '').strip()
            elif line.startswith('Prebid Info: '):
                contract['prebid_info'] = line.replace('Prebid Info: ', '').strip()
            elif line.startswith('BidNet URL: '):
                contract['bidnet_url'] = line.replace('BidNet URL: ', '').strip()
            elif line.startswith('Search Keyword: '):
                contract['search_keyword'] = line.replace('Search Keyword: ', '').strip()
            elif line.startswith('Format Type: '):
                contract['format_type'] = line.replace('Format Type: ', '').strip()
            elif line.startswith('Extracted: '):
                contract['extracted_at'] = line.replace('Extracted: ', '').strip()
        
        if 'title' in contract:
            contracts.append(contract)
    
    return contracts

def create_excel_export(contracts):
    """Create clean Excel file from contract data"""
    if not contracts:
        print("📭 No contracts to export")
        return None
    
    # Create DataFrame
    df = pd.DataFrame(contracts)
    
    # Reorder columns for better readability
    column_order = [
        'title', 'primary_agency', 'secondary_agency', 'location', 
        'description', 'prebid_info', 'bidnet_url', 'search_keyword', 
        'format_type', 'extracted_at'
    ]
    
    # Only include columns that exist in the data
    available_columns = [col for col in column_order if col in df.columns]
    df = df[available_columns]
    
    # Clean up the data
    df = df.fillna('N/A')
    df['title'] = df['title'].str.replace('\n', ' ', regex=False).str.strip()
    df['description'] = df['description'].str.replace('\n', ' ', regex=False).str.strip()
    
    # Create filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"/Users/christophernguyen/Documents/hvacscraper/smart_hvac_contracts_{timestamp}.xlsx"
    
    # Create summary statistics
    summary_stats = {
        'Metric': [
            'Total Contracts',
            'Standard Format', 
            'State Format',
            'Federal Format', 
            'Member Agency Format',
            'Unknown Format',
            'Contracts with Descriptions',
            'Contracts with Prebid Info',
            'Unique Agencies'
        ],
        'Count': [
            len(df),
            len(df[df['format_type'] == 'standard']) if 'format_type' in df else 0,
            len(df[df['format_type'] == 'state']) if 'format_type' in df else 0,
            len(df[df['format_type'] == 'federal']) if 'format_type' in df else 0,
            len(df[df['format_type'] == 'member_agency']) if 'format_type' in df else 0,
            len(df[df['format_type'] == 'unknown']) if 'format_type' in df else 0,
            len(df[df['description'] != 'N/A']) if 'description' in df else 0,
            len(df[df['prebid_info'] != 'N/A']) if 'prebid_info' in df else 0,
            len(df['primary_agency'].unique()) if 'primary_agency' in df else 0
        ]
    }
    summary_df = pd.DataFrame(summary_stats)
    
    # Write to Excel with multiple sheets
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Main contracts sheet
        df.to_excel(writer, sheet_name='Smart HVAC Contracts', index=False)
        
        # Summary statistics sheet  
        summary_df.to_excel(writer, sheet_name='Summary Statistics', index=False)
    
    # Format the Excel file
    wb = load_workbook(filename)
    
    # Format main sheet
    ws_main = wb['Smart HVAC Contracts']
    
    # Header formatting
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for cell in ws_main[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Column widths
    column_widths = {
        'A': 50,  # title
        'B': 40,  # primary_agency
        'C': 25,  # secondary_agency
        'D': 20,  # location
        'E': 60,  # description
        'F': 25,  # prebid_info
        'G': 50,  # bidnet_url
        'H': 15,  # search_keyword
        'I': 15,  # format_type
        'J': 20   # extracted_at
    }
    
    for col, width in column_widths.items():
        ws_main.column_dimensions[col].width = width
    
    # Format summary sheet
    ws_summary = wb['Summary Statistics']
    
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    
    wb.save(filename)
    
    return filename

def main():
    """Export smart parser results to Excel"""
    print("📊 Exporting Smart Contract Parser Results to Excel...")
    play_alert("Starting Excel export")
    
    # Find the latest smart contracts file
    import glob
    pattern = "/Users/christophernguyen/bidnet-hvac-scraper/data/smart_contracts_*.txt"
    files = glob.glob(pattern)
    
    if not files:
        print("❌ No smart contracts files found!")
        return
    
    # Get the most recent file
    latest_file = max(files, key=os.path.getctime)
    print(f"📖 Reading contracts from: {latest_file}")
    
    # Parse the contracts
    contracts = parse_smart_contracts_file(latest_file)
    print(f"✅ Parsed {len(contracts)} contracts")
    
    # Create Excel export
    excel_file = create_excel_export(contracts)
    
    if excel_file:
        print(f"📈 Excel file created: {excel_file}")
        print(f"📂 To open: open \"{excel_file}\"")
        play_alert(f"Excel export complete: {len(contracts)} contracts")
    else:
        print("❌ Failed to create Excel file")
        play_alert("Excel export failed")

if __name__ == "__main__":
    main()